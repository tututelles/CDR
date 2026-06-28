#!/usr/bin/env python3
"""
CDR Produtos — Servidor local de sincronização
Executa: python3 server.py
Acesse:  http://localhost:5000
"""

import json, os, threading
from flask import Flask, jsonify, request, send_from_directory
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX_PATH = os.path.join(os.path.dirname(__file__), "produtos_cdr.xlsx")
SHEET     = "Página3"
HEADERS   = ["ID", "Nome do item", "Tipo", "Coluna 1", "Estoque", "Status"]

app = Flask(__name__, static_folder=os.path.dirname(__file__))

# ── XLSX helpers ──────────────────────────────────────────────────────────────

def read_products():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET]

    # Find header row (row with "Nome do item")
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Nome do item":
                header_row = cell.row
                break
        if header_row:
            break
    if not header_row:
        return []

    col_map = {}
    for cell in ws[header_row]:
        if cell.value in HEADERS:
            col_map[cell.value] = cell.column

    products = []
    for r in range(header_row + 1, ws.max_row + 1):
        nome = ws.cell(r, col_map.get("Nome do item", 2)).value
        status = ws.cell(r, col_map.get("Status", 6)).value
        if not nome and not status:
            continue
        products.append({
            "id":      r,                  # use row number as stable ID
            "nome":    nome   or "",
            "tipo":    ws.cell(r, col_map.get("Tipo",     3)).value or "—",
            "coluna":  ws.cell(r, col_map.get("Coluna 1", 4)).value or "—",
            "estoque": int(ws.cell(r, col_map.get("Estoque", 5)).value or 0),
            "status":  status or "Esgotado",
        })
    return products


STATUS_COLORS = {
    "Em Estoque":     ("C6EFCE", "276221"),  # green
    "Quase Acabando": ("FFEB9C", "9C5700"),  # yellow
    "Esgotado":       ("FFC7CE", "9C0006"),  # red
}

def _thin():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def write_products(products):
    # Rebuild sheet from scratch keeping wb intact
    wb = openpyxl.load_workbook(XLSX_PATH)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    # ── Header row ──────────────────────────────────────────────────────────
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="4B4E6D")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(1, ci, h)
        c.font  = hdr_font
        c.fill  = hdr_fill
        c.alignment = hdr_align
        c.border = _thin()

    ws.row_dimensions[1].height = 22

    # ── Data rows ───────────────────────────────────────────────────────────
    for ri, p in enumerate(products, 2):
        bg, fg = STATUS_COLORS.get(p["status"], ("FFFFFF", "000000"))
        fill = PatternFill("solid", fgColor=bg)

        vals = [p["id"], p["nome"], p["tipo"], p["coluna"], p["estoque"], p["status"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v)
            c.font   = Font(name="Arial", size=10,
                            color=fg if ci == 6 else "000000",
                            bold=(ci == 6))
            c.fill   = fill
            c.border = _thin()
            c.alignment = Alignment(horizontal="center" if ci in (1,3,4,5,6) else "left",
                                    vertical="center")
        ws.row_dimensions[ri].height = 18

    # ── Column widths ───────────────────────────────────────────────────────
    widths = [8, 32, 14, 12, 12, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    wb.save(XLSX_PATH)


# ── Routes ────────────────────────────────────────────────────────────────────

@app.after_request
def cors(r):
    r.headers["Access-Control-Allow-Origin"]  = "*"
    r.headers["Access-Control-Allow-Headers"] = "Content-Type"
    r.headers["Access-Control-Allow-Methods"] = "GET,POST,PUT,DELETE,OPTIONS"
    return r

@app.route("/api/products", methods=["GET"])
def get_products():
    return jsonify(read_products())

@app.route("/api/products", methods=["POST"])
def add_product():
    data = request.json
    products = read_products()
    new_id = max((p["id"] for p in products), default=1) + 1
    products.append({
        "id":      new_id,
        "nome":    data.get("nome", ""),
        "tipo":    data.get("tipo", "—"),
        "coluna":  data.get("coluna", "—"),
        "estoque": int(data.get("estoque", 0)),
        "status":  data.get("status", "Em Estoque"),
    })
    write_products(products)
    return jsonify({"ok": True, "id": new_id})

@app.route("/api/products/<int:pid>", methods=["PUT"])
def update_product(pid):
    data = request.json
    products = read_products()
    for p in products:
        if p["id"] == pid:
            p.update({
                "nome":    data.get("nome",    p["nome"]),
                "tipo":    data.get("tipo",    p["tipo"]),
                "coluna":  data.get("coluna",  p["coluna"]),
                "estoque": int(data.get("estoque", p["estoque"])),
                "status":  data.get("status",  p["status"]),
            })
            break
    write_products(products)
    return jsonify({"ok": True})

@app.route("/api/products/<int:pid>", methods=["DELETE"])
def delete_product(pid):
    products = [p for p in read_products() if p["id"] != pid]
    write_products(products)
    return jsonify({"ok": True})

@app.route("/api/products", methods=["OPTIONS"])
@app.route("/api/products/<int:pid>", methods=["OPTIONS"])
def options(_=None):
    return "", 204

@app.route("/")
def index():
    return send_from_directory(os.path.dirname(__file__), "controle_produtos_cdr.html")

if __name__ == "__main__":
    print("\n╔══════════════════════════════════════╗")
    print("║  CDR Produtos — Servidor iniciado    ║")
    print("║  Acesse: http://localhost:5000       ║")
    print("╚══════════════════════════════════════╝\n")
    app.run(port=5000, debug=False)
